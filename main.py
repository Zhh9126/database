#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import warnings
# 过滤所有警告
warnings.filterwarnings("ignore")
"""
MySQL MGR集群一键安装工具
"""

import os
import sys
import json
import time
import paramiko
import subprocess
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd
from datetime import datetime
import getpass
import socket
import shutil
from dataclasses import dataclass
from enum import Enum
import logging
import tempfile
import re

# 配置日志
logging.basicConfig(
    level=logging.WARNING,  # 降低日志级别，减少输出
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(f'mysql_mgr_install_{datetime.now().strftime("%Y%m%d_%H%M%S")}.log', encoding='utf-8')
    ]
)
logger = logging.getLogger(__name__)

class MGRMode(Enum):
    """MGR模式"""
    SINGLE_PRIMARY = "single_primary"
    MULTI_PRIMARY = "multi_primary"

@dataclass
class ServerConfig:
    """服务器配置"""
    hostname: str
    ip: str
    ssh_port: int = 22
    ssh_user: str = "root"
    ssh_password: str = ""
    mysql_root_password: str = "123456"
    mysql_port: int = 3306
    mgr_port: int = 33061
    server_id: int = 1
    mgr_mode: MGRMode = MGRMode.MULTI_PRIMARY
    replication_user: str = "repl"
    replication_password: str = "repl"
    data_dir: str = "/data/mysql"
    install_dir: str = "/usr/local/mysql"
    mysql_version: str = "8.0.33"
    mysql_package: str = "mysql-8.0.33-linux-glibc2.12-x86_64.tar.xz"

class SSHExecutor:
    """SSH执行器"""
    
    def __init__(self, server_config):
        self.server = server_config
        self.ssh = None
        self.sftp = None
    
    def connect(self):
        """连接服务器"""
        try:
            self.ssh = paramiko.SSHClient()
            self.ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            self.ssh.connect(
                hostname=self.server.ip,
                port=self.server.ssh_port,
                username=self.server.ssh_user,
                password=self.server.ssh_password,
                timeout=30,
                banner_timeout=30,
                auth_timeout=30
            )
            return True
        except Exception as e:
            print(f"  ✗ 连接失败 {self.server.hostname}: {e}")
            return False
    
    def exec_command(self, command, timeout=60, show_output=False):
        """执行命令"""
        try:
            stdin, stdout, stderr = self.ssh.exec_command(command, timeout=timeout)
            exit_code = stdout.channel.recv_exit_status()
            output = stdout.read().decode('utf-8', errors='ignore').strip()
            error = stderr.read().decode('utf-8', errors='ignore').strip()
            
            return {
                'success': exit_code == 0,
                'exit_code': exit_code,
                'output': output,
                'error': error
            }
        except Exception as e:
            if show_output:
                print(f"  ✗ 执行命令失败: {command} - {e}")
            return {
                'success': False,
                'exit_code': -1,
                'output': '',
                'error': str(e)
            }
    
    def upload_file(self, local_path, remote_path):
        """上传文件"""
        try:
            if not self.sftp:
                self.sftp = self.ssh.open_sftp()
            
            remote_dir = os.path.dirname(remote_path)
            self.exec_command(f"mkdir -p {remote_dir}")
            
            self.sftp.put(local_path, remote_path)
            return True
        except Exception as e:
            print(f"  ✗ 文件上传失败: {e}")
            return False
    
    def close(self):
        """关闭连接"""
        if self.sftp:
            self.sftp.close()
        if self.ssh:
            self.ssh.close()

class MySQLMGRInstaller:
    """MySQL MGR集群安装器 - 简化版"""
    
    def __init__(self):
        self.servers = []
        self.template_file = "mysql_mgr_template.xlsx"
        self.cluster_uuid = "aaaaaaaa-aaaa-aaaa-aaaa-aaaaaaaaaaaa"
        self.mysql_package_path = None
    
    def main_menu(self):
        """主菜单"""
        while True:
            print("\n" + "="*60)
            print("MySQL MGR集群一键安装工具 - 简化版")
            print("="*60)
            print("1. 手动搭建")
            print("2. 批量搭建（使用模板文件）")
            print("3. 创建模板")
            print("4. 退出")
            print("="*60)
            
            choice = input("请选择操作 (1-4): ").strip()
            
            if choice == "1":
                self.manual_setup()
            elif choice == "2":
                self.batch_setup()
            elif choice == "3":
                self.create_template()
            elif choice == "4":
                print("感谢使用，再见！")
                sys.exit(0)
            else:
                print("无效选择，请重新输入")
    
    def manual_setup(self):
        """手动搭建"""
        print("\n" + "="*60)
        print("手动搭建MySQL MGR集群")
        print("="*60)
        
        mgr_mode = MGRMode.MULTI_PRIMARY
        mode = input("选择MGR模式 (1-单主模式, 2-多主模式, 默认2): ").strip()
        if mode == "1":
            mgr_mode = MGRMode.SINGLE_PRIMARY
        
        try:
            server_count = int(input("请输入服务器数量 (2-10, 默认3): ").strip() or "3")
            server_count = max(2, min(server_count, 10))
        except:
            server_count = 3
        
        self.servers = []
        for i in range(server_count):
            print(f"\n配置第 {i+1} 台服务器:")
            server = ServerConfig(
                hostname=input(f"主机名 (默认node{i+1}): ").strip() or f"node{i+1}",
                ip=input(f"IP地址: ").strip(),
                ssh_port=int(input("SSH端口 (默认22): ").strip() or "22"),
                ssh_user=input("SSH用户名 (默认root): ").strip() or "root",
                ssh_password=getpass.getpass(f"SSH密码: "),
                mysql_root_password=getpass.getpass(f"MySQL root密码 (默认123456): ") or "123456",
                server_id=i+1,
                mgr_mode=mgr_mode
            )
            self.servers.append(server)
        
        self.print_config_summary()
        confirm = input("\n确认以上配置开始安装? (y/n): ").strip().lower()
        
        if confirm == 'y':
            self.upload_mysql_package()
            self.install_cluster()
    
    def batch_setup(self):
        """批量搭建"""
        print("\n" + "="*60)
        print("批量搭建MySQL MGR集群")
        print("="*60)
        
        template_file = input(f"模板文件路径 (默认{self.template_file}): ").strip() or self.template_file
        
        if not os.path.exists(template_file):
            print(f"模板文件 {template_file} 不存在，请先创建模板")
            return
        
        try:
            df = pd.read_excel(template_file)
            self.servers = []
            
            for _, row in df.iterrows():
                mgr_mode_str = str(row.get('mgr_mode', 'multi_primary')).lower()
                mgr_mode = MGRMode.SINGLE_PRIMARY if mgr_mode_str == 'single_primary' else MGRMode.MULTI_PRIMARY
                
                server = ServerConfig(
                    hostname=str(row['hostname']),
                    ip=str(row['ip']),
                    ssh_port=int(row.get('ssh_port', 22)),
                    ssh_user=str(row.get('ssh_user', 'root')),
                    ssh_password=str(row.get('ssh_password', '')),
                    mysql_root_password=str(row.get('mysql_root_password', '123456')),
                    mysql_port=int(row.get('mysql_port', 3306)),
                    mgr_port=int(row.get('mgr_port', 33061)),
                    server_id=int(row.get('server_id', 1)),
                    mgr_mode=mgr_mode,
                    replication_user=str(row.get('replication_user', 'repl')),
                    replication_password=str(row.get('replication_password', 'repl')),
                    data_dir=str(row.get('data_dir', '/data/mysql')),
                    install_dir=str(row.get('install_dir', '/usr/local/mysql')),
                    mysql_version=str(row.get('mysql_version', '8.0.33')),
                    mysql_package=str(row.get('mysql_package', 'mysql-8.0.33-linux-glibc2.12-x86_64.tar.xz'))
                )
                self.servers.append(server)
            
            self.print_config_summary()
            confirm = input("\n确认以上配置开始安装? (y/n): ").strip().lower()
            
            if confirm == 'y':
                self.upload_mysql_package()
                self.install_cluster()
                
        except Exception as e:
            print(f"读取模板文件失败: {e}")
    
    def create_template(self):
        """创建模板"""
        print("\n" + "="*60)
        print("创建MySQL MGR集群配置模板")
        print("="*60)
        
        template_file = input(f"模板文件保存路径 (默认{self.template_file}): ").strip() or self.template_file
        
        wb = Workbook()
        ws = wb.active
        ws.title = "MySQL MGR配置"
        
        headers = [
            "hostname", "ip", "ssh_port", "ssh_user", "ssh_password",
            "mysql_root_password", "mysql_port", "mgr_port", "server_id",
            "mgr_mode", "replication_user", "replication_password",
            "data_dir", "install_dir", "mysql_version", "mysql_package"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        example_data = [
            ["node1", "192.168.59.249", 22, "root", "123123", "123456", 3306, 33061, 1, "multi_primary", "repl", "repl", "/data/mysql", "/usr/local/mysql", "8.0.33", "mysql-8.0.33-linux-glibc2.12-x86_64.tar.xz"],
            ["node2", "192.168.59.250", 22, "root", "123123", "123456", 3306, 33061, 2, "multi_primary", "repl", "repl", "/data/mysql", "/usr/local/mysql", "8.0.33", "mysql-8.0.33-linux-glibc2.12-x86_64.tar.xz"],
            ["node3", "192.168.59.251", 22, "root", "123123", "123456", 3306, 33061, 3, "multi_primary", "repl", "repl", "/data/mysql", "/usr/local/mysql", "8.0.33", "mysql-8.0.33-linux-glibc2.12-x86_64.tar.xz"]
        ]
        
        for row_idx, row_data in enumerate(example_data, 2):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(template_file)
        print(f"✓ 模板文件已创建: {template_file}")
        print("\n模板字段说明:")
        print("1. mgr_mode: single_primary (单主模式) 或 multi_primary (多主模式)")
        print("2. server_id: 每台服务器需要唯一的server_id")
        print("3. 根据实际情况修改IP地址、密码等配置")
    
    def print_config_summary(self):
        """打印配置摘要"""
        print("\n" + "="*60)
        print("配置摘要")
        print("="*60)
        print(f"服务器数量: {len(self.servers)}")
        print(f"MGR模式: {self.servers[0].mgr_mode.value}")
        print("\n服务器列表:")
        for i, server in enumerate(self.servers, 1):
            print(f"  节点{i}: {server.hostname} ({server.ip}) - Server ID: {server.server_id}")
    
    def upload_mysql_package(self):
        """上传MySQL安装包"""
        print("\n" + "="*60)
        print("上传MySQL安装包")
        print("="*60)
        
        default_package = "mysql-8.0.33-linux-glibc2.12-x86_64.tar.xz"
        package_path = input(f"MySQL安装包路径 (默认{default_package}): ").strip() or default_package
        
        if not os.path.exists(package_path):
            print(f"✗ 安装包 {package_path} 不存在，请确保文件存在")
            return False
        
        self.mysql_package_path = package_path
        print(f"✓ 安装包路径: {package_path}")
        return True
    
    def install_cluster(self):
        """安装集群"""
        print("\n" + "="*60)
        print("开始安装MySQL MGR集群")
        print("="*60)
        
        start_time = time.time()
        
        try:
            # 1. 准备所有节点环境
            print("\n[1/7] 准备所有节点环境...")
            for server in self.servers:
                if not self.prepare_node(server):
                    print(f"✗ 节点 {server.hostname} 环境准备失败")
                    return
            
            # 2. 在所有节点上安装MySQL
            print("\n[2/7] 安装MySQL...")
            for server in self.servers:
                if not self.install_mysql(server):
                    print(f"✗ 节点 {server.hostname} MySQL安装失败")
                    return
            
            # 3. 配置第一个节点并引导集群
            print("\n[3/7] 配置第一个节点并引导集群...")
            first_node = self.servers[0]
            if not self.configure_and_bootstrap_first_node(first_node):
                print("✗ 第一个节点配置失败")
                return
            
            # 4. 配置其他节点并加入集群
            print("\n[4/7] 配置其他节点并加入集群...")
            for server in self.servers[1:]:
                if not self.configure_and_join_node(server):
                    print(f"✗ 节点 {server.hostname} 加入集群失败")
                    return
            
            # 5. 在所有节点上创建repl用户
            print("\n[5/7] 创建复制用户...")
            for server in self.servers:
                if not self.create_replication_user(server):
                    print(f"✗ 节点 {server.hostname} 复制用户创建失败")
                    return
            
            # 6. 重新启动MGR以确保用户权限生效
            print("\n[6/7] 重新启动MGR集群...")
            if not self.restart_mgr_cluster():
                print("✗ MGR集群重启失败")
                return
            
            # 7. 验证集群状态
            print("\n[7/7] 验证集群状态...")
            self.verify_cluster()
            
            elapsed_time = time.time() - start_time
            print(f"\n" + "="*60)
            print(f"✓ MySQL MGR集群安装完成！")
            print(f"总耗时: {elapsed_time:.2f} 秒")
            print("="*60)
            
            # 显示连接信息
            self.show_connection_info()
            
        except Exception as e:
            print(f"\n✗ 安装过程中发生错误: {e}")
    
    def prepare_node(self, server):
        """准备节点环境"""
        print(f"  准备节点 {server.hostname}...", end="")
        
        ssh = SSHExecutor(server)
        if not ssh.connect():
            print(" ✗ 连接失败")
            return False
        
        try:
            # 1. 关闭SELinux和防火墙
            ssh.exec_command("setenforce 0")
            ssh.exec_command("sed -i 's/SELINUX=enforcing/SELINUX=disabled/' /etc/selinux/config")
            ssh.exec_command("systemctl stop firewalld")
            ssh.exec_command("systemctl disable firewalld")
            
            # 2. 停止现有MySQL进程
            ssh.exec_command("pkill -9 mysqld 2>/dev/null || true")
            
            # 3. 创建用户和目录
            ssh.exec_command("groupadd mysql 2>/dev/null || true")
            ssh.exec_command("useradd -r -g mysql -s /bin/false mysql 2>/dev/null || true")
            ssh.exec_command(f"mkdir -p {server.data_dir}")
            ssh.exec_command(f"mkdir -p {server.install_dir}")
            
            # 4. 清理数据目录
            ssh.exec_command(f"rm -rf {server.data_dir}/*")
            ssh.exec_command(f"chown -R mysql:mysql {server.data_dir}")
            ssh.exec_command(f"chown -R mysql:mysql {server.install_dir}")
            
            # 5. 设置主机名
            ssh.exec_command(f"hostnamectl set-hostname {server.hostname}")
            
            # 6. 上传MySQL安装包
            if self.mysql_package_path:
                package_name = os.path.basename(server.mysql_package)
                remote_path = f"/opt/{package_name}"
                ssh.upload_file(self.mysql_package_path, remote_path)
            
            ssh.close()
            print(" ✓ 完成")
            return True
            
        except Exception as e:
            ssh.close()
            print(" ✗ 失败")
            return False
    
    def install_mysql(self, server):
        """安装MySQL"""
        print(f"  在节点 {server.hostname} 上安装MySQL...", end="")
        
        ssh = SSHExecutor(server)
        if not ssh.connect():
            print(" ✗ 连接失败")
            return False
        
        try:
            package_name = os.path.basename(server.mysql_package)
            
            # 1. 解压安装包
            ssh.exec_command(f"cd /opt && tar -xvf {package_name}")
            
            # 2. 移动到安装目录
            ssh.exec_command(f"rm -rf {server.install_dir}/*")
            ssh.exec_command(f"mv /opt/mysql-{server.mysql_version}-linux-glibc2.12-x86_64/* {server.install_dir}/")
            ssh.exec_command(f"chown -R mysql:mysql {server.install_dir}")
            
            # 3. 创建配置文件
            my_cnf = self.generate_my_cnf(server)
            
            with tempfile.NamedTemporaryFile(mode='w', delete=False) as f:
                f.write(my_cnf)
                temp_config = f.name
            
            ssh.upload_file(temp_config, "/etc/my.cnf")
            os.unlink(temp_config)
            
            # 4. 初始化MySQL
            init_cmd = f"cd {server.install_dir}/bin && "
            init_cmd += f"./mysqld --defaults-file=/etc/my.cnf "
            init_cmd += f"--basedir={server.install_dir} "
            init_cmd += f"--datadir={server.data_dir} "
            init_cmd += f"--user=mysql --initialize-insecure"
            
            ssh.exec_command(init_cmd)
            
            # 5. 创建启动脚本
            self.create_mysql_service(ssh, server)
            
            # 6. 启动MySQL
            ssh.exec_command("/etc/init.d/mysql start")
            
            # 等待MySQL启动
            time.sleep(3)
            
            # 7. 设置环境变量
            #ssh.exec_command(f"echo 'export PATH=\\$PATH:{server.install_dir}/bin' >> /etc/profile")
            #ssh.exec_command(f"echo 'export MYSQL_HOME={server.install_dir}' >> /etc/profile")
            #ssh.exec_command("source /etc/profile")

# 3sh.exec_command("echo 'export PATH=/usr/local/sbin:/usr/local/bin:/usr/sbin:/usr/bin:/sbin:/bin:/root/bin:$PATH' >> /etc/profile")
            ssh.exec_command(f"echo 'export PATH=$PATH:{server.install_dir}/bin' >> /etc/profile")
            ssh.exec_command(f"echo 'export MYSQL_HOME={server.install_dir}' >> /etc/profile")
            ssh.exec_command("source /etc/profile")
            # 8. 创建MySQL命令软链接
            ssh.exec_command(f"ln -sf {server.install_dir}/bin/mysql /usr/bin/mysql 2>/dev/null || true")
            ssh.exec_command(f"ln -sf {server.install_dir}/bin/mysqladmin /usr/bin/mysqladmin 2>/dev/null || true")
            ssh.exec_command(f"ln -sf {server.install_dir}/bin/mysqldump /usr/bin/mysqldump 2>/dev/null || true")
            
            # 9. 更新hosts文件
            self.update_hosts_file(ssh)
            
            # 10. 检查MySQL进程
            result = ssh.exec_command("ps aux | grep mysqld | grep -v grep")
            
            ssh.close()
            
            if "mysqld" in result['output']:
                print(" ✓ 完成")
                return True
            else:
                print(" ✗ 失败")
                return False
            
        except Exception as e:
            ssh.close()
            print(" ✗ 失败")
            return False
    
    def update_hosts_file(self, ssh):
        """更新hosts文件"""
        hosts_content = """127.0.0.1   localhost localhost.localdomain localhost4 localhost4.localdomain4
::1         localhost localhost.localdomain localhost6 localhost6.localdomain6
"""
        
        for server in self.servers:
            hosts_content += f"{server.ip}    {server.hostname}\n"
        
        with tempfile.NamedTemporaryFile(mode='w', delete=False) as f:
            f.write(hosts_content)
            temp_file = f.name
        
        ssh.upload_file(temp_file, "/etc/hosts")
        os.unlink(temp_file)
    
    def generate_my_cnf(self, server):
        """生成my.cnf配置文件"""
        seeds = ','.join([f"{s.ip}:{s.mgr_port}" for s in self.servers])
        ip_list = ','.join([s.ip for s in self.servers])
        
        config = f"""[client]
port={server.mysql_port}
socket=/tmp/mysql.sock

[mysqld]
# Basic settings
port={server.mysql_port}
socket=/tmp/mysql.sock
basedir={server.install_dir}
datadir={server.data_dir}
pid-file={server.data_dir}/mysql.pid
log-error={server.data_dir}/mysql.err

# Character set
character-set-server=utf8mb4
collation-server=utf8mb4_unicode_ci

# Server ID
server-id={server.server_id}

# Binary logging
log-bin={server.data_dir}/mysql-bin
binlog_format=ROW
sync_binlog=1
binlog_expire_logs_seconds=604800
max_binlog_size=100M

# Replication
relay-log={server.data_dir}/relay-bin
relay-log-index={server.data_dir}/relay-bin.index

# GTID
gtid_mode=ON
enforce_gtid_consistency=ON

# InnoDB
innodb_buffer_pool_size=128M
innodb_log_file_size=48M
innodb_log_buffer_size=16M
innodb_flush_log_at_trx_commit=1
innodb_lock_wait_timeout=50
innodb_file_per_table=ON
innodb_autoinc_lock_mode=2

# SQL mode for MySQL 8.0
sql_mode=STRICT_TRANS_TABLES,NO_ZERO_IN_DATE,NO_ZERO_DATE,ERROR_FOR_DIVISION_BY_ZERO,NO_ENGINE_SUBSTITUTION

# MGR Settings
disabled_storage_engines="MyISAM,BLACKHOLE,FEDERATED,ARCHIVE,MEMORY"
loose-group_replication_group_name="{self.cluster_uuid}"
loose-group_replication_start_on_boot=OFF
loose-group_replication_bootstrap_group=OFF
loose-group_replication_local_address="{server.ip}:{server.mgr_port}"
loose-group_replication_group_seeds="{seeds}"
loose-group_replication_ip_whitelist="{ip_list}"
loose-group_replication_single_primary_mode={'OFF' if server.mgr_mode == MGRMode.MULTI_PRIMARY else 'ON'}
loose-group_replication_enforce_update_everywhere_checks={'ON' if server.mgr_mode == MGRMode.MULTI_PRIMARY else 'OFF'}

# Other settings
skip_name_resolve=1
explicit_defaults_for_timestamp=true
log_timestamps=SYSTEM
lower_case_table_names=1

[mysql]
prompt='\\u@\\h [\\d]> '
default-character-set=utf8mb4
"""
        
        return config
    
    def create_mysql_service(self, ssh, server):
        """创建MySQL启动脚本"""
        service_content = f"""#!/bin/bash
#
# mysql        Start and stop the mysql database server daemon
#
basedir={server.install_dir}
datadir={server.data_dir}

start(){{
    echo -n "Starting MySQL: "
    nohup $basedir/bin/mysqld --defaults-file=/etc/my.cnf --user=mysql >/dev/null 2>&1 &
    ret=$?
    sleep 3
    [ $ret -eq 0 ] && echo "OK" || echo "FAILED"
    return $ret
}}

stop(){{
    echo -n "Stopping MySQL: "
    $basedir/bin/mysqladmin -uroot shutdown 2>/dev/null || pkill -9 mysqld 2>/dev/null
    ret=$?
    sleep 2
    [ $ret -eq 0 ] && echo "OK" || echo "FAILED"
    return $ret
}}

restart(){{
    stop
    sleep 2
    start
}}

case "$1" in
    start)
        start
        ;;
    stop)
        stop
        ;;
    restart)
        restart
        ;;
    *)
        echo "Usage: $0 {{start|stop|restart}}"
        exit 2
esac

exit $?
"""
        
        with tempfile.NamedTemporaryFile(mode='w', delete=False) as f:
            f.write(service_content)
            temp_service = f.name
        
        ssh.upload_file(temp_service, "/etc/init.d/mysql")
        os.unlink(temp_service)
        
        ssh.exec_command("chmod +x /etc/init.d/mysql")
        ssh.exec_command("chkconfig --add mysql 2>/dev/null || true")
        ssh.exec_command("chkconfig mysql on 2>/dev/null || true")
    
    def configure_and_bootstrap_first_node(self, server):
        """配置并引导第一个节点"""
        print(f"  配置节点 {server.hostname} 并引导集群...", end="")
        
        ssh = SSHExecutor(server)
        if not ssh.connect():
            print(" ✗ 连接失败")
            return False
        
        try:
            time.sleep(5)
            
            # 1. 设置root密码
            mysql_cmd = f"{server.install_dir}/bin/mysql -S /tmp/mysql.sock"
            
            for i in range(5):
                result = ssh.exec_command(f"{mysql_cmd} -e \"SELECT 1;\"")
                if result['success']:
                    break
                time.sleep(3)
            
            set_pw_cmd = f"{mysql_cmd} -e \"ALTER USER 'root'@'localhost' IDENTIFIED BY '{server.mysql_root_password}';\""
            result = ssh.exec_command(set_pw_cmd)
            
            if not result['success']:
                print(" ✗ 设置密码失败")
                return False
            
            # 2. 创建远程root用户
            mysql_cmd_pw = f"{server.install_dir}/bin/mysql -p{server.mysql_root_password} -S /tmp/mysql.sock"
            
            cmds = [
                f"{mysql_cmd_pw} -e \"CREATE USER IF NOT EXISTS 'root'@'%' IDENTIFIED BY '{server.mysql_root_password}';\"",
                f"{mysql_cmd_pw} -e \"GRANT ALL PRIVILEGES ON *.* TO 'root'@'%' WITH GRANT OPTION;\"",
                f"{mysql_cmd_pw} -e \"FLUSH PRIVILEGES;\""
            ]
            
            for cmd in cmds:
                ssh.exec_command(cmd)
            
            # 3. 配置MGR（先不创建repl用户）
            mgr_cmds = [
                f"{mysql_cmd_pw} -e \"INSTALL PLUGIN group_replication SONAME 'group_replication.so';\"",
                f"{mysql_cmd_pw} -e \"SET GLOBAL group_replication_bootstrap_group=ON;\"",
                f"{mysql_cmd_pw} -e \"START GROUP_REPLICATION;\"",
                f"{mysql_cmd_pw} -e \"SET GLOBAL group_replication_bootstrap_group=OFF;\""
            ]
            
            for cmd in mgr_cmds:
                ssh.exec_command(cmd)
            
            # 4. 检查集群状态
            time.sleep(3)
            check_cmd = f"{mysql_cmd_pw} -e \"SELECT MEMBER_HOST, MEMBER_PORT, MEMBER_STATE FROM performance_schema.replication_group_members;\""
            result = ssh.exec_command(check_cmd)
            
            ssh.close()
            
            if "ONLINE" in result['output']:
                print(" ✓ 完成")
                return True
            else:
                print(" ✗ 失败")
                return False
            
        except Exception as e:
            ssh.close()
            print(" ✗ 失败")
            return False
    
    def configure_and_join_node(self, server):
        """配置节点并加入集群"""
        print(f"  配置节点 {server.hostname} 并加入集群...", end="")
        
        ssh = SSHExecutor(server)
        if not ssh.connect():
            print(" ✗ 连接失败")
            return False
        
        try:
            time.sleep(5)
            
            # 1. 设置root密码
            mysql_cmd = f"{server.install_dir}/bin/mysql -S /tmp/mysql.sock"
            
            for i in range(5):
                result = ssh.exec_command(f"{mysql_cmd} -e \"SELECT 1;\"")
                if result['success']:
                    break
                time.sleep(3)
            
            set_pw_cmd = f"{mysql_cmd} -e \"ALTER USER 'root'@'localhost' IDENTIFIED BY '{server.mysql_root_password}';\""
            ssh.exec_command(set_pw_cmd)
            
            # 2. 创建远程root用户
            mysql_cmd_pw = f"{server.install_dir}/bin/mysql -p{server.mysql_root_password} -S /tmp/mysql.sock"
            
            cmds = [
                f"{mysql_cmd_pw} -e \"CREATE USER IF NOT EXISTS 'root'@'%' IDENTIFIED BY '{server.mysql_root_password}';\"",
                f"{mysql_cmd_pw} -e \"GRANT ALL PRIVILEGES ON *.* TO 'root'@'%' WITH GRANT OPTION;\"",
                f"{mysql_cmd_pw} -e \"FLUSH PRIVILEGES;\""
            ]
            
            for cmd in cmds:
                ssh.exec_command(cmd)
            
            # 3. 配置MGR（先不创建repl用户）
            mgr_cmds = [
                f"{mysql_cmd_pw} -e \"INSTALL PLUGIN group_replication SONAME 'group_replication.so';\"",
                f"{mysql_cmd_pw} -e \"SET GLOBAL group_replication_bootstrap_group=OFF;\"",
                f"{mysql_cmd_pw} -e \"RESET MASTER;\""
            ]
            
            for cmd in mgr_cmds:
                ssh.exec_command(cmd)
            
            # 4. 加入集群
            join_cmd = f"{mysql_cmd_pw} -e \"START GROUP_REPLICATION;\""
            result = ssh.exec_command(join_cmd)
            
            if not result['success']:
                # 尝试修复
                fix_cmds = [
                    f"{mysql_cmd_pw} -e \"STOP GROUP_REPLICATION;\"",
                    f"{mysql_cmd_pw} -e \"RESET MASTER;\"",
                    f"{mysql_cmd_pw} -e \"START GROUP_REPLICATION;\""
                ]
                
                for cmd in fix_cmds:
                    ssh.exec_command(cmd)
            
            # 5. 检查集群状态
            time.sleep(5)
            check_cmd = f"{mysql_cmd_pw} -e \"SELECT MEMBER_HOST, MEMBER_PORT, MEMBER_STATE FROM performance_schema.replication_group_members;\""
            result = ssh.exec_command(check_cmd)
            
            ssh.close()
            
            if "ONLINE" in result['output'] or "RECOVERING" in result['output']:
                print(" ✓ 完成")
                return True
            else:
                print(" ✗ 失败")
                return False
            
        except Exception as e:
            ssh.close()
            print(" ✗ 失败")
            return False
    
    def create_replication_user(self, server):
        """创建复制用户（在所有节点上分别创建）"""
        print(f"  在节点 {server.hostname} 上创建复制用户...", end="")
        
        ssh = SSHExecutor(server)
        if not ssh.connect():
            print(" ✗ 连接失败")
            return False
        
        try:
            mysql_cmd_pw = f"{server.install_dir}/bin/mysql -p{server.mysql_root_password} -S /tmp/mysql.sock"
            
            # 1. 创建复制用户（使用不同的SID避免二进制日志冲突）
            create_user_cmds = [
                f"{mysql_cmd_pw} -e \"SET SQL_LOG_BIN=0;\"",
                f"{mysql_cmd_pw} -e \"DROP USER IF EXISTS '{server.replication_user}'@'%';\"",
                f"{mysql_cmd_pw} -e \"CREATE USER '{server.replication_user}'@'%' IDENTIFIED WITH mysql_native_password BY '{server.replication_password}';\"",
                f"{mysql_cmd_pw} -e \"GRANT REPLICATION SLAVE ON *.* TO '{server.replication_user}'@'%';\"",
                f"{mysql_cmd_pw} -e \"GRANT BACKUP_ADMIN ON *.* TO '{server.replication_user}'@'%';\"",
                f"{mysql_cmd_pw} -e \"GRANT GROUP_REPLICATION_STREAM ON *.* TO '{server.replication_user}'@'%';\"",
                f"{mysql_cmd_pw} -e \"GRANT CREATE USER ON *.* TO '{server.replication_user}'@'%';\"",
                f"{mysql_cmd_pw} -e \"FLUSH PRIVILEGES;\"",
                f"{mysql_cmd_pw} -e \"SET SQL_LOG_BIN=1;\""
            ]
            
            for cmd in create_user_cmds:
                ssh.exec_command(cmd)
            
            # 2. 配置group_replication_recovery通道
            recovery_cmds = [
                f"{mysql_cmd_pw} -e \"STOP GROUP_REPLICATION;\"",
                f"{mysql_cmd_pw} -e \"CHANGE REPLICATION SOURCE TO SOURCE_USER='{server.replication_user}', SOURCE_PASSWORD='{server.replication_password}' FOR CHANNEL 'group_replication_recovery';\"",
                f"{mysql_cmd_pw} -e \"START GROUP_REPLICATION;\""
            ]
            
            for cmd in recovery_cmds:
                ssh.exec_command(cmd)
            
            ssh.close()
            print(" ✓ 完成")
            return True
            
        except Exception as e:
            ssh.close()
            print(" ✗ 失败")
            return False
    
    def restart_mgr_cluster(self):
        """重新启动MGR集群以确保用户权限生效"""
        print("  重新启动MGR集群...", end="")
        
        try:
            # 首先停止所有节点的MGR
            for server in self.servers:
                ssh = SSHExecutor(server)
                if ssh.connect():
                    mysql_cmd_pw = f"{server.install_dir}/bin/mysql -p{server.mysql_root_password} -S /tmp/mysql.sock"
                    ssh.exec_command(f"{mysql_cmd_pw} -e \"STOP GROUP_REPLICATION;\"")
                    ssh.close()
                time.sleep(1)
            
            # 等待所有节点停止
            time.sleep(3)
            
            # 重新启动第一个节点引导集群
            first_node = self.servers[0]
            ssh = SSHExecutor(first_node)
            if ssh.connect():
                mysql_cmd_pw = f"{first_node.install_dir}/bin/mysql -p{first_node.mysql_root_password} -S /tmp/mysql.sock"
                ssh.exec_command(f"{mysql_cmd_pw} -e \"SET GLOBAL group_replication_bootstrap_group=ON;\"")
                ssh.exec_command(f"{mysql_cmd_pw} -e \"START GROUP_REPLICATION;\"")
                ssh.exec_command(f"{mysql_cmd_pw} -e \"SET GLOBAL group_replication_bootstrap_group=OFF;\"")
                ssh.close()
            
            # 等待第一个节点启动
            time.sleep(5)
            
            # 其他节点重新加入
            for server in self.servers[1:]:
                ssh = SSHExecutor(server)
                if ssh.connect():
                    mysql_cmd_pw = f"{server.install_dir}/bin/mysql -p{server.mysql_root_password} -S /tmp/mysql.sock"
                    ssh.exec_command(f"{mysql_cmd_pw} -e \"START GROUP_REPLICATION;\"")
                    ssh.close()
                time.sleep(2)
            
            print(" ✓ 完成")
            return True
            
        except Exception as e:
            print(" ✗ 失败")
            return False
    
    def verify_cluster(self):
        """验证集群状态"""
        print("\n验证集群状态:")
        
        online_count = 0
        for server in self.servers:
            print(f"  检查节点 {server.hostname}...", end="")
            ssh = SSHExecutor(server)
            if not ssh.connect():
                print(" ✗ 连接失败")
                continue
            
            try:
                mysql_cmd = f"{server.install_dir}/bin/mysql -p{server.mysql_root_password} -S /tmp/mysql.sock"
                check_cmd = f"{mysql_cmd} -e \"SELECT MEMBER_HOST, MEMBER_PORT, MEMBER_STATE FROM performance_schema.replication_group_members ORDER BY MEMBER_HOST;\""
                result = ssh.exec_command(check_cmd)
                
                ssh.close()
                
                if result['success']:
                    if "ONLINE" in result['output']:
                        online_count += 1
                        print(" ✓ 在线")
                    else:
                        print(" ✗ 离线")
                else:
                    print(" ✗ 无法获取状态")
                
            except Exception as e:
                print(" ✗ 检查失败")
        
        print(f"\n集群状态: {online_count}/{len(self.servers)} 个节点在线")
        if online_count == len(self.servers):
            print("✓ 集群状态正常")
        else:
            print("⚠ 集群状态异常")
    
    def show_connection_info(self):
        """显示连接信息"""
        print("\n" + "="*60)
        print("连接信息:")
        print("="*60)
        
        first_server = self.servers[0]
        print(f"集群管理节点: {first_server.hostname} ({first_server.ip})")
        print(f"MySQL端口: {first_server.mysql_port}")
        print(f"MySQL root密码: {first_server.mysql_root_password}")
        print(f"复制用户: {first_server.replication_user}")
        print(f"复制密码: {first_server.replication_password}")
        print("\n连接到MySQL的命令:")
        print(f"  mysql -h {first_server.ip} -P {first_server.mysql_port} -uroot -p{first_server.mysql_root_password}")
        
        print("\n各节点信息:")
        for server in self.servers:
            print(f"  {server.hostname} ({server.ip}):")
            print(f"    MySQL端口: {server.mysql_port}")
            print(f"    MGR端口: {server.mgr_port}")
            print(f"    Server ID: {server.server_id}")
        
        print("\n环境变量已设置，可在任意目录使用mysql命令:")
        print("  mysql -uroot -p  # 本地连接")
        print("  mysql -h <host> -P <port> -uroot -p  # 远程连接")
        print("="*60)

def main():
    """主函数"""
    try:
        import paramiko
        import pandas as pd
        import openpyxl
    except ImportError as e:
        print(f"缺少必要模块: {e}")
        print("请安装以下模块:")
        print("pip install paramiko pandas openpyxl")
        sys.exit(1)
    
    print("="*60)
   # print("MySQL MGR集群一键安装工具 - 简化版")
    print("="*60)
    
    installer = MySQLMGRInstaller()
    installer.main_menu()

if __name__ == "__main__":
    main()
